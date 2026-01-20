"""
validator.py

All the "real logic" lives here so app.py stays very small.

This module contains:
- PDF parsing + rule extraction (conservative heuristics)
- Excel parsing (openpyxl) into a normalized structure
- Validation engine that produces explainable issues

We keep the MVP simple and conservative:
- Only validate numerical-code column (NOT narrative text).
- Treat unclear / low-confidence rules as warnings-only.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, asdict
from typing import Dict, List, Optional, Tuple, Any

import openpyxl

import unicodedata
# PDF parsing: try pdfplumber if installed (best), else fall back to PyPDF2.
try:
    import pdfplumber  # type: ignore
except Exception:
    pdfplumber = None

try:
    from PyPDF2 import PdfReader  # type: ignore
except Exception:
    PdfReader = None


# ----------------------------
# Data structures
# ----------------------------

@dataclass
class Rule:
    """
    A conservative, explainable validation rule extracted from the PDF.
    """
    rule_id: str
    source: str  # "pdf"
    confidence: str  # "high" | "medium" | "low"
    kind: str  # "allowed_values" | "conditional_required" | "conditional_na"
    variables: List[str]  # variables involved (with brackets, e.g. "[educ_post_expand]")
    when: Optional[Dict[str, Any]]  # condition in structured-ish form (see builder funcs)
    assert_: Dict[str, Any]  # what must hold if condition is true
    else_: Optional[Dict[str, Any]]  # what must hold if condition is false
    evidence: str  # snippet for explainability

@dataclass
class Issue:
    severity: str  # "error" | "warning"
    tab_name: str
    variable: str
    message: str
    expected: str
    actual: str
    evidence: str

@dataclass
class TabSummary:
    errors: int
    warnings: int
    ok: int

# ----------------------------
# PDF extraction
# ----------------------------

VAR_PATTERN = re.compile(r"\[[A-Za-z0-9_]+\]")

def _read_pdf_text(pdf_path: str) -> str:
    """
    Extract text from a PDF file.

    This is intentionally simple. For MVP v1:
    - We only need enough text to find variable names and a few simple patterns.
    """
    if pdfplumber is not None:
        text_parts: List[str] = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                text_parts.append(txt)
        return "\n".join(text_parts)

    if PdfReader is None:
        raise RuntimeError("No PDF parser available. Install pdfplumber or PyPDF2.")

    reader = PdfReader(pdf_path)
    text_parts = []
    for page in reader.pages:
        try:
            text_parts.append(page.extract_text() or "")
        except Exception:
            text_parts.append("")
    return "\n".join(text_parts)


def extract_pdf_model(pdf_path: str) -> Dict[str, Any]:
    """
    Extract a 'PDF model' from the uploaded codebook PDF.

    Output:
    {
      "variables": set([...]),
      "rules": [Rule as dict, ...],
      "summary": {...}
    }
    """
    text = _read_pdf_text(pdf_path)
    variables = sorted(set(VAR_PATTERN.findall(text)))

    rules: List[Rule] = []
    rules.extend(_extract_yes_no_rules(text))
    rules.extend(_extract_enumeration_rules(text))
    rules.extend(_extract_land_expro_comp_rule(text))
    rules.extend(_extract_restrict_cleav_and_year_rules(text))

    summary = {
        "variables_detected": len(variables),
        "rules_extracted": len(rules),
        "rules_by_confidence": {
            "high": sum(1 for r in rules if r.confidence == "high"),
            "medium": sum(1 for r in rules if r.confidence == "medium"),
            "low": sum(1 for r in rules if r.confidence == "low"),
        }
    }

    return {
        "variables": variables,
        "rules": [asdict(r) for r in rules],
        "summary": summary,
        "raw_text_sample": "\n".join(text.splitlines()[:120])  # helpful for debugging
    }


def _extract_yes_no_rules(text: str) -> List[Rule]:
    """
    Enforce allowed values {0,1} for variables marked as (Yes = 1, No = 0).

    Handles:
      A) [var] line followed by a wrapped continuation line containing (Yes = 1, No = 0)
      B) a header line containing (Yes = 1, No = 0) followed by a dash-list of [vars]
    """
    rules: List[Rule] = []
    lines = text.splitlines()

    def add(var: str, evidence: str) -> None:
        rules.append(Rule(
            rule_id=f"YESNO_{var.strip('[]')}",
            source="pdf",
            confidence="high",
            kind="allowed_values",
            variables=[var],
            when=None,
            assert_={"allowed_values": [0, 1]},
            else_=None,
            evidence=evidence
        ))

    # ---- Pattern A: [var] + nearby yes/no, but stop if next line introduces a different var ----
    for i, line in enumerate(lines):
        vars_in_line = VAR_PATTERN.findall(line)
        if not vars_in_line:
            continue

        # Collect a short "continuation" blob until a *different* variable appears
        cont = [line]
        for k in range(1, 3):  # look ahead up to 2 lines
            if i + k >= len(lines):
                break
            nxt = lines[i + k]
            nxt_vars = VAR_PATTERN.findall(nxt)
            if nxt_vars and any(nv not in vars_in_line for nv in nxt_vars):
                break
            cont.append(nxt)

        blob = " ".join(cont)
        if "Yes = 1" in blob and "No = 0" in blob:
            evidence = " ".join([c.strip() for c in cont if c.strip()][:2])
            for var in vars_in_line:
                add(var, evidence or line.strip())
            continue

        # Common wrap: yes/no on the immediate next line
        if i + 1 < len(lines):
            nxt = lines[i + 1]
            if ("Yes = 1" in nxt and "No = 0" in nxt) and not VAR_PATTERN.search(nxt):
                for var in vars_in_line:
                    add(var, f"{line.strip()} {nxt.strip()}")

    # ---- Pattern B: header line with Yes/No applies to a following dash-list of variables ----
    i = 0
    while i < len(lines):
        line = lines[i]
        if ("Yes = 1" in line and "No = 0" in line) and not VAR_PATTERN.search(line):
            evidence = line.strip()
            collected: List[str] = []

            for j in range(i + 1, min(i + 35, len(lines))):
                nxt = lines[j].strip()
                if not nxt:
                    if collected:
                        break
                    continue
                if nxt.startswith("•"):  # next question block
                    break
                if nxt.startswith(("– [", "- [", "— [")):
                    collected.extend(VAR_PATTERN.findall(nxt))
                    continue
                if collected and VAR_PATTERN.search(nxt):
                    break

            for var in sorted(set(collected)):
                add(var, evidence)

        i += 1

    # De-dup by variable
    dedup: Dict[str, Rule] = {}
    for r in rules:
        k = r.variables[0]
        if k not in dedup:
            dedup[k] = r
    return list(dedup.values())


def _extract_enumeration_rules(text: str) -> List[Rule]:
    """
    Heuristic: detect enumerated code options like:
      1. No restrictions
      2. Restricted access
      3. Restricted actions

    IMPORTANT: Avoid false positives where the PDF has numbered *variable lists* like:
      1. [health_post_vacc]
      2. [health_post_dis]
    Those are NOT enumerated value options.
    """
    rules: List[Rule] = []
    lines = text.splitlines()

    i = 0
    while i < len(lines):
        line = lines[i]
        vars_in_line = VAR_PATTERN.findall(line)
        if not vars_in_line:
            i += 1
            continue

        # ✅ Only start enumerations from bullet-style variable declarations ("• [var] ...")
        # This avoids misreading numbered variable lists as allowed values.
        if not line.lstrip().startswith("•"):
            i += 1
            continue

        var = vars_in_line[0]
        allowed: List[int] = []
        evidence_lines = [line.strip()]

        for j in range(i + 1, min(i + 26, len(lines))):
            nxt = lines[j].strip()

            # Stop if we hit the next variable/bullet.
            if nxt.startswith("• ["):
                break

            # ✅ If a numbered line contains a different bracketed var, it's not an option list.
            if re.match(r"^\s*\d+\.\s*", lines[j]) and VAR_PATTERN.search(nxt) and (var not in nxt):
                break

            m = re.match(r"^\s*(\d+)\.\s*", lines[j])
            if m:
                allowed.append(int(m.group(1)))
                evidence_lines.append(nxt)
            else:
                if allowed:
                    evidence_lines.append(nxt)

        if len(allowed) >= 2:
            rules.append(Rule(
                rule_id=f"ENUM_{var.strip('[]')}",
                source="pdf",
                confidence="medium",
                kind="allowed_values",
                variables=[var],
                when=None,
                assert_={"allowed_values": sorted(set(allowed))},
                else_=None,
                evidence="\n".join([l for l in evidence_lines if l][:10])
            ))

        i += 1

    return rules


def _extract_land_expro_comp_rule(text: str) -> List[Rule]:
    """
    Special-case rule present in the codebook text:
      [land_post_expro_comp] If Yes to any of the above...

    "Any of the above" refers to the preceding expropriation questions:
      [land_post_expro_dev]
      [land_post_expro_redist]
      [land_post_expro_corr]
      [land_post_expro_other]

    This is *medium confidence* because it depends on PDF ordering, but in this codebook
    it's quite stable.
    """
    # Only add this rule if the key variable appears in the PDF.
    if "[land_post_expro_comp]" not in text:
        return []

    evidence_snippet = ""
    for line in text.splitlines():
        if "[land_post_expro_comp]" in line:
            evidence_snippet = line.strip()
            break

    parent_vars = [
        "[land_post_expro_dev]",
        "[land_post_expro_redist]",
        "[land_post_expro_corr]",
        "[land_post_expro_other]",
    ]

    return [Rule(
        rule_id="LAND_EXPRO_COMP_REQUIRED",
        source="pdf",
        confidence="medium",
        kind="conditional_required",
        variables=parent_vars + ["[land_post_expro_comp]"],
        when={
            "any_equals": [{"var": v, "value": 1} for v in parent_vars],
        },
        assert_={
            "var": "[land_post_expro_comp]",
            "required": True,  # must not be blank/NA
        },
        else_={
            "var": "[land_post_expro_comp]",
            "must_be_blank": True,  # should be blank when condition is false
        },
        evidence=evidence_snippet or "If Yes to any of the above. Were landowners compensated..."
    )]


def _extract_restrict_cleav_and_year_rules(text: str) -> List[Rule]:
    """
    Special-case rules:
      [educ_post_restrict_cleav] If restricted (else write NA) ...
      [educ_post_restrict_year] If restricted: What year(s) ...

    In the template, both are in the Education sheet.

    We interpret "restricted" using:
      [educ_post_restrict_group] numeric code:
        1 = no new restrictions
        2/3 = restricted

    This is medium confidence: it's based on the explicit code options.
    """
    rules: List[Rule] = []
    if "[educ_post_restrict_group]" not in text:
        return rules

    def add_rule(child_var: str, rule_id: str, evidence_contains: str) -> None:
        if child_var not in text:
            return
        evidence = ""
        for line in text.splitlines():
            if child_var in line:
                evidence = line.strip()
                break
        rules.append(Rule(
            rule_id=rule_id,
            source="pdf",
            confidence="medium",
            kind="conditional_na",
            variables=["[educ_post_restrict_group]", child_var],
            when={
                "var": "[educ_post_restrict_group]",
                "in": [2, 3],  # restricted
            },
            assert_={
                "var": child_var,
                "required": True,  # must have a code when restricted
            },
            else_={
                "var": child_var,
                "must_be_na": True,  # else write NA
            },
            evidence=evidence or evidence_contains
        ))

    add_rule("[educ_post_restrict_cleav]", "EDUC_POST_RESTRICT_CLEAV", "If restricted (else write NA)")
    add_rule("[educ_post_restrict_year]", "EDUC_POST_RESTRICT_YEAR", "If restricted")

    return rules


_MONTHS = {m.lower() for m in [
    "January","February","March","April","May","June","July","August",
    "September","October","November","December"
]}
_CITATION_STOPSTART = {
    # to avoid false positives like "(signed on 20th June 1999)"
    "signed","accessed","see","as","on","in","during","from","at","by","for",
    "with","without","after","before","between","since","until","updated","retrieved"
}

def _norm_name(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = s.replace("’", "'").replace("–", "-").replace("—", "-")
    s = re.sub(r"[^A-Za-z0-9\-'\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def _extract_reference_index(wb: "openpyxl.Workbook") -> Dict[str, Any]:
    """
    Build a set of normalized author/org tokens from the References/Citations tab.
    Handles newline-separated references in a single cell.
    Handles em-dash repeated-author lines (e.g., '———. 2013. ...').
    """
    name_map = {n.lower(): n for n in wb.sheetnames}
    ref_sheet = None
    for key in ["references", "citations", "bibliography", "sources"]:
        if key in name_map:
            ref_sheet = name_map[key]
            break

    if not ref_sheet:
        return {"sheet": None, "names_norm": set(), "raw_lines": 0}

    ws = wb[ref_sheet]

    # Collect all text from column B (like your template)
    raw_cells: List[str] = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if not val:
            continue
        s = str(val).strip()
        if not s or s.lower() == "references":
            continue
        raw_cells.append(s)

    lines: List[str] = []
    for cell_text in raw_cells:
        for ln in cell_text.splitlines():
            ln = ln.strip()
            if ln:
                lines.append(ln)

    names_norm: set = set()
    prev_author_block: Optional[str] = None

    def add_phrase_variants(author_block: str) -> None:
        toks = [t for t in author_block.split() if t]
        if len(toks) < 2:
            return
        # Only add org-like phrases; skip obvious multi-author lists
        if re.search(r"\b(and|&)\b", author_block) or author_block.count(",") >= 2:
            return
        for n in (2, 3, 4):
            if len(toks) >= n:
                names_norm.add(_norm_name(" ".join(toks[:n])))
        names_norm.add(_norm_name(" ".join(toks[-2:])))

    for ln in lines:
        # If line begins with em-dash / repeated author, substitute previous author block
        if re.match(r"^[—–-]{2,}", ln) and prev_author_block:
            rest = re.sub(r"^[—–-]{2,}\s*", "", ln).lstrip(". ").strip()
            ln = f"{prev_author_block}. {rest}"

        ym = re.search(r"\b(\d{4}[a-z]?|n\.d\.)\b", ln)
        if not ym:
            continue

        author_block = ln[:ym.start()].strip().rstrip(".")
        if author_block:
            prev_author_block = author_block
            # Always include the full author/org block to catch multi-word orgs
            names_norm.add(_norm_name(author_block))
            add_phrase_variants(author_block)

        # If commas exist, likely "Lastname, Firstname, and Lastname ..."
        if "," in author_block:
            tmp = author_block.replace(" and ", ", ").replace(" & ", ", ")
            parts = [p.strip() for p in tmp.split(",") if p.strip()]
            for p in parts:
                w = p.split()
                if not w:
                    continue
                if w[0].lower() == "and" and len(w) > 1:
                    w = w[1:]
                # heuristic: surname sometimes appears first (comma style) OR last (e.g., "Egbert Sondorp")
                candidates = {w[0], w[-1]}
                for c in candidates:
                    if re.match(r"^[A-Z]", c):
                        names_norm.add(_norm_name(c))
        else:
            # org / single-name author: keep full phrase plus first/last token
            full = author_block.strip()
            if full:
                names_norm.add(_norm_name(full))
                toks = full.split()
                if toks:
                    names_norm.add(_norm_name(toks[0]))
                    names_norm.add(_norm_name(toks[-1]))

    return {"sheet": ref_sheet, "names_norm": names_norm, "raw_lines": len(lines)}

def _extract_cited_names_from_narrative(narrative: Any) -> List[Dict[str, Any]]:
    """
    Extract likely citation tokens from narrative parentheses that contain a year or n.d.
    Avoid false positives like "(signed on 20th June 1999)".
    """
    
    def _remove_quoted_spans(text: str) -> str:
        """
        Remove content inside straight or curly quotes to avoid flagging citations that appear
        inside quoted source text.
        """
        # Handles: "..."  '...'  “...”  ‘...’
        return re.sub(r'(["“”\'])(?:(?=(\\?))\2.)*?\1', "", text)

    if not narrative:
        return []

    out: List[Dict[str, Any]] = []
    text = _remove_quoted_spans(str(narrative))

    for par in re.findall(r"\(([^()]{0,250})\)", text):
        for seg in par.split(";"):
            seg = seg.strip()
            ym = re.search(r"\b(\d{4}[a-z]?|n\.d\.)\b", seg)
            if not ym:
                continue

            author_part = seg[:ym.start()].strip()
            if not author_part:
                continue

            first = re.match(r"^\s*([A-Za-z]+)", author_part)
            if first and first.group(1).lower() in _CITATION_STOPSTART:
                continue

            author_part = re.sub(r"\bet al\.?\b", "", author_part)

            toks = re.findall(r"[A-Z][A-Za-zÀ-ÖØ-öø-ÿ'’-]+", author_part)
            toks = [t for t in toks if t.lower() not in _MONTHS]
            if not toks:
                continue

            phrases = []
            if len(toks) >= 2:
                phrases.append(" ".join(toks[:2]))  # handles org-ish "Medica Kosova"

            out.append({"tokens": toks, "phrases": phrases, "raw": seg})

    return out
# ----------------------------
# Excel parsing
# ----------------------------

def _normalize_code_value(value: Any) -> Optional[Any]:
    """
    Convert the "Numerical code" cell into a normalized form.

    We accept:
    - int / float numeric codes
    - strings like "1: Yes" -> 1
    - strings like "NA" -> "NA"
    - empty -> None
    """
    if value is None:
        return None

    # Some spreadsheets store numbers as floats (e.g. 1.0). Convert safely.
    if isinstance(value, (int, float)):
        # Treat NaN as empty
        try:
            if value != value:  # NaN check
                return None
        except Exception:
            pass
        return int(value)

    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        if s.upper() == "NA":
            return "NA"
        # "1: Yes" or "1 - Yes"
        m = re.match(r"^\s*(\d+)\s*[:\-]", s)
        if m:
            return int(m.group(1))
        # If it's a pure integer string
        if re.fullmatch(r"\d+", s):
            return int(s)
        return s  # keep as-is (might be text codes)
    return value


# --- Excel template color-code helpers ---

def _cell_has_fill(cell) -> bool:
    """True if the cell has a background fill (used for template color-coding)."""
    try:
        f = cell.fill
        return bool(getattr(f, "patternType", None)) and f.patternType != "none"
    except Exception:
        return False


def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _has_any_year(v: Any) -> bool:
    """Loose check: contains at least one 4-digit year (or a single numeric year)."""
    if v is None:
        return False

    if isinstance(v, (int, float)):
        try:
            if v != v:  # NaN
                return False
        except Exception:
            pass
        y = int(v)
        return 1000 <= y <= 2200

    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.upper() == "NA":
            return False
        return re.search(r"\b(1[0-9]{3}|20[0-9]{2}|21[0-9]{2})\b", s) is not None

    return False


def parse_excel_workbook(xlsx_path: str) -> Dict[str, Any]:
    """
    Parse the Excel workbook into a normalized structure.

    Output:
    {
      "path": ...,
      "tabs": {
        "Education": {
            "conflict_title": "Kosovo War (1998 - 1999)",
            "rows": {
              "[educ_pre_restrict_quota]": {
                 "code": 1,
                 "year": 1990,
                 "policy": ...,
                 "narrative": ...   # present but not validated
              }, ...
            }
        }, ...
      }
    }

    We assume the template structure:
      Column B: Variable
      Column C: Numerical code
      Column D: Year(s) implemented
      Column E: Formal policies or laws
      Column F: Narrative
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    reference_index = _extract_reference_index(wb)
    tabs: Dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        conflict_title = ws["B2"].value  # in sample template and sample Kosovo
        conflict_title = str(conflict_title).strip() if conflict_title is not None else ""

        rows: Dict[str, Dict[str, Any]] = {}

        # Variables start at row 6 in the provided template.
        for r in range(6, ws.max_row + 1):
            var_cell = ws.cell(r, 2).value  # column B
            if var_cell is None:
                continue
            var = str(var_cell).strip()
            if not var.startswith("[") or not var.endswith("]"):
                # We only treat bracketed variable IDs as variables.
                continue

            code_cell = ws.cell(r, 3)  # col C
            year_cell = ws.cell(r, 4)  # col D
            policy_cell = ws.cell(r, 5)  # col E
            narrative_cell = ws.cell(r, 6)  # col F

            code = _normalize_code_value(code_cell.value)
            year = year_cell.value  # (we don't over-normalize years)
            policy = policy_cell.value
            narrative = narrative_cell.value

            rows[var] = {
                "code": code,
                "year": year,
                "policy": policy,
                "narrative": narrative,
                "row_index": r,

                # Template-driven validation flags
                "year_required_by_color": _cell_has_fill(year_cell),
                "no_code_by_color": _cell_has_fill(code_cell),

                # Cell references for explainability
                "cell_code": f"C{r}",
                "cell_year": f"D{r}",
                "cell_policy": f"E{r}",
            }

        tabs[sheet_name] = {
            "conflict_title": conflict_title,
            "rows": rows,
        }

    return {"path": xlsx_path, "tabs": tabs, "reference_index": reference_index}


# ----------------------------
# Validation engine
# ----------------------------

PLACEHOLDER_CONFLICT = "CONFLICT (Start - End)"

def validate_workbook(pdf_model: Dict[str, Any], workbook: Dict[str, Any]) -> Dict[str, Any]:
    """
    Validate the parsed workbook against:
    - structural checks from PRD
    - high/medium confidence rules extracted from the PDF
    - citation cross-check: cited author/org in narrative must exist in References tab

    Returns a report dict that can be rendered as HTML or JSON.
    """
    pdf_vars = set(pdf_model.get("variables", []))
    rules = [Rule(**r) for r in pdf_model.get("rules", [])]

    # --- Workbook-level issues (NOT tied to any single tab) ---
    workbook_issues: List[Issue] = []

    # --- Citation reference index (built during parse_excel_workbook) ---
    ref_index = workbook.get("reference_index", {}) or {}
    ref_sheet = ref_index.get("sheet")
    ref_names = ref_index.get("names_norm", set()) or set()
    citations_enabled = bool(ref_sheet and ref_names)

    # If citation checking isn't possible, record a single workbook-level warning once
    if not citations_enabled:
        workbook_issues.append(Issue(
            severity="warning",
            tab_name="(workbook)",
            variable="(references)",
            message="No References/Citations sheet detected (or it was empty); citation cross-check skipped.",
            expected="A sheet named 'References' (or 'Citations') with reference entries in column B.",
            actual=repr(ref_sheet),
            evidence="Add a References tab or ensure references are present in column B."
        ))

    all_tabs_report: List[Dict[str, Any]] = []
    total_errors = 0
    total_warnings = 0

    for tab_name, tab in workbook["tabs"].items():
        if ref_sheet and tab_name == ref_sheet:
            continue # don't validate the references sheet as a coding tab
        issues: List[Issue] = []
        ok_checks = 0

        # 1) Structural: conflict title in B2
        ct = tab.get("conflict_title", "").strip()
        if ct == "" or ct == PLACEHOLDER_CONFLICT:
            issues.append(Issue(
                severity="warning",
                tab_name=tab_name,
                variable="(metadata)",
                message="Missing conflict title in cell B2 (expected a conflict name like 'Kosovo War (1998 - 1999)').",
                expected="A non-placeholder conflict title in B2",
                actual=repr(ct),
                evidence="Template expects the conflict title in cell B2."
            ))
        else:
            ok_checks += 1

        # 2) Structural: unrecognized variables
        for var in tab["rows"].keys():
            if var not in pdf_vars:
                issues.append(Issue(
                    severity="warning",
                    tab_name=tab_name,
                    variable=var,
                    message="Variable exists in Excel but was not found in the PDF codebook text (might be a mismatch/typo).",
                    expected=f"Variable should appear in codebook PDF as {var}",
                    actual=var,
                    evidence="This warning is conservative: it only checks whether the bracketed ID appears anywhere in the PDF text."
                ))
            else:
                ok_checks += 1

        # 3) Apply extracted rules (but only if their variables exist in this tab)
        for rule in rules:
            if not any(v in tab["rows"] for v in rule.variables):
                continue

            new_issues, new_ok = _apply_rule(rule, tab_name, tab["rows"])
            issues.extend(new_issues)
            ok_checks += new_ok

        # 3b) Template color-code validation (from Excel instructions)
        for var, row in tab.get("rows", {}).items():
            code = row.get("code")
            year = row.get("year")
            policy = row.get("policy")

            # Gray code cell => no numerical code allowed
            if row.get("no_code_by_color"):
                # allow blank/None and literal NA
                if code not in (None, "NA") and not _is_blank(code):
                    issues.append(Issue(
                        severity="error",
                        tab_name=tab_name,
                        variable=var,
                        message="Numerical code is not allowed (colored code cell) but a value was provided.",
                        expected="Blank (or NA if you use NA for not-applicable).",
                        actual=repr(code),
                        evidence=f"Template color-coding indicates {row.get('cell_code','Code cell')} should not contain a numerical code."
                    ))
                else:
                    ok_checks += 1

            # If policies column is filled, years column must be filled
            if not _is_blank(policy) and not _has_any_year(year):
                issues.append(Issue(
                    severity="error",
                    tab_name=tab_name,
                    variable=var,
                    message="Policy is filled out, so year must also be filled out.",
                    expected="Year(s) implemented present when policy is provided.",
                    actual=f"policy={repr(policy)}, year={repr(year)}",
                    evidence=f"Template instruction: if {row.get('cell_policy','Policy cell')} is filled, {row.get('cell_year','Year cell')} must be too."
                ))
            else:
                ok_checks += 1

        # 4) Citation cross-check (warn only)
        if citations_enabled:
            for var, row in tab.get("rows", {}).items():
                narr = row.get("narrative")
                cited = _extract_cited_names_from_narrative(narr)
                if not cited:
                    continue

                for c in cited:
                    ok = False

                    # phrase match for org-like refs (e.g., "Medica Kosova")
                    for ph in c.get("phrases", []):
                        if _norm_name(ph) in ref_names:
                            ok = True
                            break

                    # token match for surnames (e.g., Percival, Sondorp)
                    bad = []
                    for t in c.get("tokens", []):
                        if _norm_name(t) in ref_names:
                            ok = True
                        else:
                            bad.append(t)

                    if not ok and bad:
                        issues.append(Issue(
                            severity="warning",
                            tab_name=tab_name,
                            variable=var,
                            message=f"Citation author/org not found in References: {', '.join(sorted(set(bad)))}",
                            expected="Author/org name appears somewhere in the References tab.",
                            actual=f"In narrative citation: ({c.get('raw','')})",
                            evidence=f"Narrative is on row {row.get('row_index')} (column F). References sheet detected: {ref_sheet}."
                        ))

        # Summaries
        errors = sum(1 for i in issues if i.severity == "error")
        warnings = sum(1 for i in issues if i.severity == "warning")
        total_errors += errors
        total_warnings += warnings

        all_tabs_report.append({
            "tab_name": tab_name,
            "summary": asdict(TabSummary(errors=errors, warnings=warnings, ok=ok_checks)),
            "issues": [asdict(i) for i in issues],
        })

    return {
        "pdf_summary": pdf_model.get("summary", {}),
        "workbook_summary": {
            "tabs_validated": len(workbook["tabs"]),
            "total_errors": total_errors,
            "total_warnings": total_warnings,
        },
        "workbook_issues": [asdict(i) for i in workbook_issues],
        "tabs": all_tabs_report,
    }

def _apply_rule(rule: Rule, tab_name: str, rows: Dict[str, Dict[str, Any]]) -> Tuple[List[Issue], int]:
    """
    Apply a single rule to a single tab.

    Returns:
      - list of issues (errors/warnings)
      - count of "ok checks" (for a simple success metric)
    """
    issues: List[Issue] = []
    ok = 0

    if rule.kind == "allowed_values":
        var = rule.variables[0]
        if var not in rows:
            return issues, ok

        val = rows[var]["code"]
        if val is None or val == "NA":
            # We allow blank/NA because blanks are used for unfulfilled conditionals.
            # If a variable is truly required, a conditional_required rule should catch it.
            return issues, ok

        allowed = set(rule.assert_.get("allowed_values", []))
        if isinstance(val, int) and val in allowed:
            ok += 1
        else:
            issues.append(Issue(
                severity="error" if rule.confidence in ("high", "medium") else "warning",
                tab_name=tab_name,
                variable=var,
                message="Invalid numerical code (not in allowed values from codebook).",
                expected=f"One of {sorted(allowed)}",
                actual=repr(val),
                evidence=rule.evidence
            ))
        return issues, ok

    if rule.kind in ("conditional_required", "conditional_na"):
        # Evaluate condition
        cond_true, cond_unknown = _eval_condition(rule.when, rows)

        target = rule.assert_.get("var")
        if not target or target not in rows:
            return issues, ok

        actual = rows[target]["code"]

        # If we can't evaluate the condition (missing parent values), we only warn
        # when something looks suspicious.
        if cond_unknown:
            # If child has a value, warn that we couldn't verify.
            if actual not in (None, "NA"):
                issues.append(Issue(
                    severity="warning",
                    tab_name=tab_name,
                    variable=target,
                    message="Child variable is filled but parent condition could not be evaluated (missing/blank parent codes).",
                    expected="Either fill parent condition variables or leave this blank/NA until the condition is known.",
                    actual=repr(actual),
                    evidence=rule.evidence
                ))
            return issues, ok

        if cond_true:
            # Condition true: enforce rule.assert_
            if rule.assert_.get("required", False):
                if actual is None or actual == "NA":
                    issues.append(Issue(
                        severity="error" if rule.confidence in ("high", "medium") else "warning",
                        tab_name=tab_name,
                        variable=target,
                        message="Required field is missing because its parent condition is true.",
                        expected="A non-blank numerical code (not NA)",
                        actual=repr(actual),
                        evidence=rule.evidence
                    ))
                else:
                    ok += 1
        else:
            # Condition false: enforce rule.else_
            if rule.else_:
                else_var = rule.else_.get("var", target)
                else_actual = rows[else_var]["code"] if else_var in rows else None

                if rule.else_.get("must_be_blank", False):
                    if else_actual is None:
                        ok += 1
                    elif else_actual == "NA":
                        issues.append(Issue(
                            severity="warning",
                            tab_name=tab_name,
                            variable=else_var,
                            message="Field is NA, but codebook flow prefers leaving it blank when the condition is false.",
                            expected="Blank",
                            actual="NA",
                            evidence=rule.evidence
                        ))
                    else:
                        issues.append(Issue(
                            severity="error" if rule.confidence in ("high", "medium") else "warning",
                            tab_name=tab_name,
                            variable=else_var,
                            message="Field should be blank because parent condition is false, but it is filled.",
                            expected="Blank",
                            actual=repr(else_actual),
                            evidence=rule.evidence
                        ))

                if rule.else_.get("must_be_na", False):
                    if else_actual == "NA":
                        ok += 1
                    elif else_actual is None:
                        # PRD: warn if NA required but blank used
                        issues.append(Issue(
                            severity="warning",
                            tab_name=tab_name,
                            variable=else_var,
                            message="Codebook says to write NA when condition is false, but the cell is blank.",
                            expected="NA",
                            actual="Blank",
                            evidence=rule.evidence
                        ))
                    else:
                        issues.append(Issue(
                            severity="error" if rule.confidence in ("high", "medium") else "warning",
                            tab_name=tab_name,
                            variable=else_var,
                            message="Codebook says to write NA when condition is false, but the cell is filled with another value.",
                            expected="NA",
                            actual=repr(else_actual),
                            evidence=rule.evidence
                        ))
        return issues, ok

    # Unknown rule kind: ignore in MVP
    return issues, ok


def _eval_condition(when: Optional[Dict[str, Any]], rows: Dict[str, Dict[str, Any]]) -> Tuple[bool, bool]:
    """
    Evaluate a very small "condition language".

    Returns (condition_is_true, condition_is_unknown)

    Supported forms:
    - {"any_equals": [{"var": "[x]", "value": 1}, ...]}
    - {"var": "[x]", "in": [2,3]}
    """
    if not when:
        return True, False

    # any_equals
    if "any_equals" in when:
        any_list = when["any_equals"]
        unknown = False
        for item in any_list:
            var = item["var"]
            want = item["value"]
            if var not in rows:
                unknown = True
                continue
            val = rows[var]["code"]
            if val is None or val == "NA":
                unknown = True
                continue
            if val == want:
                return True, False
        return False, unknown

    # var in [..]
    if "var" in when and "in" in when:
        var = when["var"]
        allowed = set(when["in"])
        if var not in rows:
            return False, True
        val = rows[var]["code"]
        if val is None:
            return False, True
        if val == "NA":
            return False, True
        if isinstance(val, int) and val in allowed:
            return True, False
        return False, False

    # Unknown condition type -> treat as unknown to be conservative
    return False, True
